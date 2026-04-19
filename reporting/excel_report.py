from datetime import datetime
import os
import shutil
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import BarChart, Reference


# ---------------- UTILITIES ----------------

def auto_adjust_columns(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + 3


def style_headers(ws):
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")


def add_bar_chart(ws, title, data_col, category_col, position):
    chart = BarChart()
    chart.title = title

    data = Reference(ws, min_col=data_col, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=category_col, min_row=2, max_row=ws.max_row)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    ws.add_chart(chart, position)


# ---------------- MAIN FUNCTION ----------------

def generate_excel_report(kpis, analysis, anomalies, insights):
    os.makedirs("reports", exist_ok=True)

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    file_path = f"reports/financial_report_{timestamp}.xlsx"
    latest_path = "reports/latest_report.xlsx"

    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:

        # -------- KPIs --------
        pd.DataFrame([kpis]).to_excel(writer, sheet_name="KPIs", index=False)

        # -------- Core Analysis --------
        analysis["monthly"].to_excel(writer, sheet_name="Monthly", index=False)
        analysis["category"].to_excel(writer, sheet_name="Category", index=False)
        analysis["region"].to_excel(writer, sheet_name="Region", index=False)

        # -------- Additional Analysis --------
        if "company" in analysis:
            analysis["company"].to_excel(writer, sheet_name="Company", index=False)

        if "risk_summary" in analysis:
            analysis["risk_summary"].to_excel(writer, sheet_name="Risk", index=False)

        # -------- Anomalies --------
        if anomalies is not None:
            anomalies.to_excel(writer, sheet_name="Anomalies", index=False)

        # -------- Insights --------
        pd.DataFrame({"Insights": insights}).to_excel(
            writer, sheet_name="Insights", index=False
        )

    # ---------------- POST FORMATTING ----------------
    wb = load_workbook(file_path)

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        auto_adjust_columns(ws)
        style_headers(ws)

    # -------- Title for KPIs --------
    ws = wb["KPIs"]
    ws.insert_rows(1)
    ws["A1"] = "Automated Financial Analysis Report"
    ws["A1"].font = Font(size=14, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    # -------- Highlight Anomalies --------
    if "Anomalies" in wb.sheetnames:
        ws = wb["Anomalies"]
        red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

        headers = [cell.value for cell in ws[1]]

        if "expense_ratio" in headers:
            idx = headers.index("expense_ratio")

            for row in ws.iter_rows(min_row=2):
                val = row[idx].value
                if val and val > 1.5:
                    for cell in row:
                        cell.fill = red_fill

    # ---------------- ADD CHARTS ----------------

    # Monthly Revenue
    if "Monthly" in wb.sheetnames:
        ws = wb["Monthly"]
        add_bar_chart(ws, "Monthly Revenue", data_col=3, category_col=2, position="H2")

    # Category Expense
    if "Category" in wb.sheetnames:
        ws = wb["Category"]
        add_bar_chart(ws, "Expense by Category", data_col=3, category_col=1, position="H2")

    # Region Profit
    if "Region" in wb.sheetnames:
        ws = wb["Region"]
        add_bar_chart(ws, "Profit by Region", data_col=3, category_col=1, position="H2")

    # Company Profit
    if "Company" in wb.sheetnames:
        ws = wb["Company"]
        add_bar_chart(ws, "Company Profit Comparison", data_col=3, category_col=1, position="H2")

    # ---------------- FORMAT INSIGHTS ----------------
    ws = wb["Insights"]
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    # Save file
    wb.save(file_path)

    # Save latest version
    shutil.copy(file_path, latest_path)

    print(f"[INFO] Report saved: {file_path}")
    print("[INFO] Latest report updated")